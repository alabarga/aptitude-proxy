from django.shortcuts import render
from django.views.generic import *
from django.urls import reverse, reverse_lazy
from django.http import HttpResponseRedirect
from django.contrib.auth.models import User
from aptitude_crd.models import *
from django.db.models import Max

from django.contrib.auth.mixins import LoginRequiredMixin

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from rest_framework import viewsets
from django.shortcuts import get_object_or_404

from .forms import EvaluacionForm, SearchForm, RegistroForm, ScreeningForm
import pandas as pd
from io import BytesIO as IO
from django.http import HttpResponse
import json

from datetime import datetime
from django.utils.translation import gettext as _
from django.contrib.auth import get_user_model

from aptitude_crd.models import _INTERVENCION
from django.contrib.auth.forms import UserCreationForm
from registration.backends.admin_approval.views import RegistrationView
from django.contrib.sites.shortcuts import get_current_site
from registration import signals

from django.shortcuts import redirect

from io import BytesIO, StringIO
from openpyxl import Workbook

UserModel = get_user_model

class HomeView(TemplateView):
    template_name = "aptitude/home.html"

class AvisoLegal(TemplateView):
    template_name = "legal/aviso_legal.html"

class ProteccionDatos(TemplateView):
    template_name = "legal/proteccion_datos.html"

class TerminosCondiciones(TemplateView):
    template_name = "legal/terminos-condiciones.html"

class Descargas(TemplateView):
    template_name = "aptitude_crd/descargas.html"

class RegistroView2(TemplateView):
    template_name = "aptitude_crd/registro.html"

class ThanksView(TemplateView):
    template_name = "aptitude_crd/gracias.html"

class IcopeView(TemplateView):
    template_name = "aptitude_crd/icope.html"

class EvalDemoView(TemplateView):
    template_name = "aptitude_crd/evaluacion_demo.html"

class MaterialesView(TemplateView):
    template_name = "aptitude_crd/materiales.html"

# class PacienteSalud(TemplateView):
#     template_name = "aptitude_crd/paciente_salud.html"

class AptitudeRegistrationView(RegistrationView):
    def register(self, form):

        site = get_current_site(self.request)

        if hasattr(form, 'save'):
            new_user_instance = form.save(commit=False)
        else:
            new_user_instance = (UserModel().objects
                                 .create_user(**form.cleaned_data))

        new_user = self.registration_profile.objects.create_inactive_user(
            new_user=new_user_instance,
            site=site,
            send_email=self.SEND_ACTIVATION_EMAIL,
            request=self.request,
        )
        signals.user_registered.send(sender=self.__class__,
                                     user=new_user,
                                     request=self.request)

        new_user.refresh_from_db()

        new_user.profile.lugar_trabajo = 'Navarrabiomed' #
        new_user.profile.cargo = self.request.POST['cargo']
        new_user.profile.save()

        return new_user


def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.refresh_from_db()  # load the profile instance created by the signal
            # user.profile.birth_date = form.cleaned_data.get('birth_date')
            user.save()
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=user.username, password=raw_password)
            login(request, user)
            return redirect('home')
    else:
        form = UserCreationForm()
    return render(request, 'signup.html', {'form': form})

# REGISTRO

class RegistroView(CreateView):
    model = Registro
#    fields = '__all__'
    template_name = "aptitude_crd/registro.html"
    success_url = '/gracias/'
    form_class = RegistroForm

    def get_context_data(self, **kwargs):
        context = super(RegistroView, self).get_context_data(**kwargs)
        temas = Tema.objects.all()
        context['temas'] = temas
        return context

# EVALUACIONES


# Nuevo paciente /pacientes/nuevo/

class NuevoPaciente(CreateView):
    model = Paciente
    template_name = 'aptitude_crd/paciente_nuevo.html'
    fields = '__all__'

    def get_success_url(self):

        return reverse('visita-lista',args=(self.object.id,))
    
    def form_valid(self, form):
        user = self.request.user
        form.instance.user = user
 
        return super(NuevoPaciente, self).form_valid(form)
        
def create_visit_view(request, pk):
    # Perform some action
    user = request.user
    paciente = Paciente.objects.get(pk=pk)
    visita = Visita.objects.create(paciente=paciente, 
                                    user=user)

    # Redirect to another URL, passing parameters
    return redirect('visita-detalle', pk=visita.pk)

def create_visit_step_1(request, pk):
    # Perform some action
    user = request.user
    paciente = Paciente.objects.get(pk=pk)
    visita = Visita.objects.create(paciente=paciente, 
                                    user=user,
                                    motivo = 1)

    # Redirect to another URL, passing parameters
    return redirect('visita-eval-step-1', pk=visita.pk)  
  
# Lista pacientes /pacientes/

class PacienteList(ListView):
    model = Paciente

# Buscar paciente /pacientes/buscar/

def get_or_create_paciente(apellidos, fecha_nac, dni=None):
    # Convert fecha_nac_str from 'dd/mm/yyyy' to a Python date object
    # fecha_nac = datetime.strptime(fecha_nac_str, '%d/%m/%Y').date()

    # First, try to find the patient by codigo if provided
    if dni: # First, try to find the patient by DNI if provided
        paciente, created = Paciente.objects.get_or_create(dni=dni, defaults={'apellidos': apellidos, 'fecha_nacimiento': fecha_nac})
    else:
        # Otherwise, try to find the patient by apellidos and fecha_nac
        paciente, created = Paciente.objects.get_or_create(apellidos=apellidos, fecha_nacimiento=fecha_nac)

    if created:
        print('Creating new user')
    return paciente, created

class SearchView(LoginRequiredMixin, FormView):
    template_name = 'aptitude_crd/search.html'
    form_class = SearchForm

    def form_valid(self, form):
        apellidos = form.cleaned_data['apellidos']
        fecha_nac = form.cleaned_data['fecha_nac']
        dni = form.cleaned_data['dni'].upper()
        codigo = form.cleaned_data['codigo'].upper()

        if codigo:
            patient = Paciente.objects.get(codigo=codigo)
            created = False
        else:
            patient, created = get_or_create_paciente(apellidos, fecha_nac, dni)
            
        if created:
            
            patient.codigo = "NAV{:04d}".format(patient.pk)
            patient.user = self.request.user
            patient.save()
            return HttpResponseRedirect(reverse_lazy('paciente-detalle', kwargs={'pk': patient.pk}))
        else:
            return HttpResponseRedirect(reverse_lazy('paciente-encontrado', kwargs={'pk': patient.pk}))

# Buscar paciente /pacientes/buscar/<pk>/
class PacienteCheck(DetailView):
    model = Paciente
    template_name = 'aptitude_crd/paciente_check.html'

# Buscar paciente /pacientes/<pk>/
class PacienteUpdate(UpdateView):
    model = Paciente
    fields = ['consentimiento_informado', 'sexo', 'fecha_nacimiento', 'nombre', 'apellidos', 'dni', 'ciudad', 'telefono', 'vivienda', 'convivientes', 'nivel_estudios'] # 

    def form_valid(self, form):
        print("Form is valid")
        # Access the cleaned form data and print it
        print("Form cleaned data:", form.cleaned_data)

        print("Saving form...")
        self.object = form.save()
        print("Form saved successfully")

        return super().form_valid(form)

    def form_invalid(self, form):
        print("Form is invalid")
        print(form.errors)  # Print the validation errors for debugging
        return super().form_invalid(form)

    def get_success_url(self):

        # url_host = self.request.META.get("HTTP_HOST")

        # if url_host.startswith('cat'):
        #     return reverse('paciente-salud',args=(self.object.id,))

        return reverse('visita-lista', args=(self.object.id,))

    # def get_context_data(self, **kwargs):

    #     context = super(PacienteUpdate, self).get_context_data(**kwargs)

    #     paciente = Paciente.objects.get(pk=self.kwargs.get('pk'))
    #     visitas = paciente.visitas
    #     context['visitas'] = visitas
    #     return context

# Buscar paciente /pacientes/<pk>/
class PacienteSalud(UpdateView):
    model = Paciente
    fields = ['problemas', 'problemas_num', 'salud']
    template_name = 'aptitude_crd/paciente_salud.html'

    def get_success_url(self):

        return reverse('visita-nueva',args=(self.object.id,))

    def get_context_data(self, **kwargs):

        context = super(PacienteSalud, self).get_context_data(**kwargs)

        paciente = Paciente.objects.get(pk=self.kwargs.get('pk'))
        visitas = paciente.visitas
        context['visitas'] = visitas
        return context

class PacienteDetail(DetailView):
    model = Paciente

    def get_context_data(self, **kwargs):
        context = super(PacienteDetail, self).get_context_data(**kwargs)
        paciente = Paciente.objects.get(pk=self.kwargs.get('pk'))
        visitas = paciente.visitas
        context['visitas'] = visitas
        return context

class PacienteVisitas(DetailView):
    model = Paciente
    template_name = 'aptitude_crd/paciente_visitas.html'
    def get_context_data(self, **kwargs):
        context = super(PacienteVisitas, self).get_context_data(**kwargs)
        paciente = Paciente.objects.get(pk=self.kwargs.get('pk'))
        visitas = paciente.visitas.all().order_by('-fecha')

        step1_finished = Evaluacion.objects.filter(visita__paciente__id=self.object.id).filter(completada=True).count()
        step1_pending = Evaluacion.objects.filter(visita__paciente__id=self.object.id).filter(completada=False).count()
        
        context['step1_finished'] = step1_finished
        context['step1_pending'] = step1_pending
        context['visitas'] = visitas
        return context
    


class NuevaVisita(CreateView):
    model = Visita
    fields = ['motivo', 'muestra', ]

    def get_context_data(self, **kwargs):

        context = super(NuevaVisita, self).get_context_data(**kwargs)

        paciente = get_object_or_404(Paciente, pk=self.kwargs.get('pk'))
        visitas_previas = Visita.objects.filter(paciente=paciente).filter(muestra=True)

        context['muestra_previa'] = len(visitas_previas) > 0
        return context

    def get_success_url(self):
        return reverse('visita-detalle',args=(self.object.id,))

    def form_valid(self, form):
        user = self.request.user
        paciente = get_object_or_404(Paciente, pk=self.kwargs.get('pk'))

        form.instance.user = user
        form.instance.paciente = paciente

        last_sample_code = Visita.objects.aggregate(Max('codigo_muestra'))['codigo_muestra__max'] or '00000'
        next_sample_code = '00000{}'.format(int(last_sample_code) + 1)[-5:]
        if form.instance.muestra:
            form.instance.codigo_muestra = next_sample_code

        return super(NuevaVisita, self).form_valid(form)

class VisitaDetail(DetailView):
    model = Visita

    # def get_success_url(self):
    #     screening = Evaluacion.objects.filter(visita=self.object).filter(cuestionario__order=1).first()
    #     return reverse('screening',args=(screening.id,))

    def get_context_data(self, **kwargs):

        context = super(VisitaDetail, self).get_context_data(**kwargs)

        # eval = [e for e in v.evaluaciones.all()][-1]
        # [(p.id, p.componente.nombre, p.componente.id) for p in eval.preguntas.all()]
        #
        eval_antropo = Evaluacion.objects.filter(visita=self.object).filter(cuestionario__order=0).first()
        variables = dict([('c{}'.format(p.componente.id), {'nombre': 'pregunta_{}'.format(p.id), 'valor': p.valor}) for p in eval_antropo.preguntas.all()])

        cuestionarios = Evaluacion.objects.filter(visita=self.object)
        context['cuestionarios'] = cuestionarios
        context['formularios'] = []
        for c in cuestionarios:
            context['formularios'].append(EvaluacionForm(instance=c))

        context['variables_url'] = '/evaluacion/{}/'.format(eval_antropo.id)
        context['variables'] = variables

        screening = Evaluacion.objects.filter(visita=self.object).filter(cuestionario__dominio='SCREENING').first()
        context['screening_url'] = '/evaluacion/{}/'.format(screening.id)

        return context

class ScreeningView(LoginRequiredMixin, UpdateView):
    model = Evaluacion
    form_class = ScreeningForm
    template_name = "aptitude_crd/visita_screen.html"


    def get_success_url(self):
        if len(self.object.visita.dominios_afectados()):
            return reverse('visita-eval', args=(self.object.visita.id,))
        elif self.object.visita.is_cat():
            return reverse('visita-eval', args=(self.object.visita.id,))
        else:
            return reverse('visita-cerrar', args=(self.object.visita.id,))

    def get_context_data(self, **kwargs):

        context = super(ScreeningView, self).get_context_data(**kwargs)
        evaluacion = Evaluacion.objects.get(pk=self.kwargs.get('pk'))
        context['evaluacion'] = evaluacion

        return context

    def form_valid(self, form):
        self.object = form.save()
        return super().form_valid(form)


class VisitaEval1(DetailView):
    model = Visita
    template_name = 'aptitude_crd/visita_eval.html'

    def get_context_data(self, **kwargs):

        context = super(VisitaEval1, self).get_context_data(**kwargs)

        context['primera_visita'] = True
        context['dominios'] = ['ANTROPOMETRICO', 'SCREENING']
        
        return context


class VisitaEval(DetailView):
    model = Visita
    template_name = 'aptitude_crd/visita_eval.html'

    def get_context_data(self, **kwargs):

        context = super(VisitaEval, self).get_context_data(**kwargs)

        screening = Evaluacion.objects.filter(visita=self.object).filter(cuestionario__dominio='SCREENING').first()
        context['screening_url'] = '/screening/{}/'.format(screening.id)

        context['sensorial'] = self.object.solo_sensorial()

        if self.object.solo_sensorial() and not self.object.is_cat():
            context['dominios'] = self.object.dominios_afectados()
        else:
            context['dominios'] = self.object.lista_dominios()
        return context

class EvalView(UpdateView):
    model = Evaluacion
    form_class = EvaluacionForm
    template_name = "aptitude_crd/evaluacion_view.html"

    def get_template_names(self):
        if self.object.cuestionario.dominio == 'SCREENING': 
            return 'aptitude_crd/evaluacion_table.html'
        else:
            return 'aptitude_crd/evaluacion_view.html'

    def post(self, request, *args, **kwargs):
        # Print the POST data

        # Call the parent class to handle the form submission
        return super().post(request, *args, **kwargs)
        
    def get(self, request, *args, **kwargs):

        return super().get(request, *args, **kwargs)
    
    def get_success_url(self):
        # referer = self.request.META.get('HTTP_REFERER', reverse_lazy('home')) 
        # return referer
    
        # if self.object.cuestionario.dominio == 'ANTROPOMETRICO':
        #     # screening = Evaluacion.objects.filter(visita=self.object.visita).filter(cuestionario__order=1).first()
        #     # return reverse('screening',args=(screening.id,))
        #     return reverse('visita-detalle',args=(self.object.visita.id,))
        # else:
        #     return reverse('visita-eval',args=(self.object.visita.id,))
        
        
        # if self.object.cuestionario.dominio == 'ANTROPOMETRICO':
        #     # screening = Evaluacion.objects.filter(visita=self.object.visita).filter(cuestionario__order=1).first()
        #     # return reverse('screening',args=(screening.id,))
        #     return reverse('visita-detalle',args=(self.object.visita.id,))
        # else:
        #     return reverse('visita-eval',args=(self.object.visita.id,))

        #return reverse('visita-lista',args=(self.object.visita.paciente.id,))

        if self.object.visita.motivo == 1:
            return reverse('visita-eval-step-1',args=(self.object.visita.id,))
        else:
            return reverse('visita-eval-step-2',args=(self.object.visita.id,))

    def get_context_data(self, **kwargs):

        context = super(EvalView, self).get_context_data(**kwargs)
        evaluacion = Evaluacion.objects.get(pk=self.kwargs.get('pk'))
        context['evaluacion'] = evaluacion

        print('Recuperando contexto')
        context = super(EvalView, self).get_context_data(**kwargs)
        print(self.object.cuestionario.dominio)
        # screening = Evaluacion.objects.filter(visita=self.object).filter(cuestionario__dominio='ANTROPOMETRICO').first()
        if self.object.cuestionario.pk == 13:
            print('Creando contexto')
            id_height = self.object.preguntas.get(componente__id = 120).id
            id_weight = self.object.preguntas.get(componente__id = 121).id
            id_imc = self.object.preguntas.get(componente__id = 122).id

            context['id_height'] = id_height
            context['id_weight'] = id_weight
            context['id_imc'] = id_imc
            print(context)
        return context

class VisitaCierre(UpdateView):
    model = Visita
    fields = ['dominios', 'intervencion', 'notas']
    template_name = 'aptitude_crd/visita_cierre.html'

    def get_context_data(self, **kwargs):
        context = super(VisitaCierre, self).get_context_data(**kwargs)

        context['dominios_afectados'] = len(self.object.dominios_afectados()) > 0
        context['dominios'] = ['SENSORIAL', 'FUNCIONAL', 'VITALIDAD', 'COGNITIVO', 'PSICOLOGICO']
        context['tipo_intervencion'] = [ {'short': p[0], 'long': p[1] } for p in _INTERVENCION]
        context['intervenciones'] = Intervencion.objects.all()
        return context

    def get_success_url(self):
        return reverse('visita-detalle',args=(self.object.id,))
        #return reverse('home')


    # def post(self, request, *args, **kwargs):
    #     form = EvaluacionForm(data=request.POST)
    #     if form.is_valid():
    #         self.send_mail(form.cleaned_data)
    #         form = EvaluacionForm()
    #         return render(request, 'aptitude/evaluacion_view.html', {'form': form})
    #     return render(request, 'aptitude/evaluacion_view.html', {'form': form})

    # def get_initial(self):
    #     initial = super(EvalView, self).get_initial()
    #     if self.request.user.is_authenticated:
    #         initial.update({'name': self.request.user.get_full_name()})
    #     return initial

    # def form_valid(self, form):
    #     self.send_mail(form.cleaned_data)
    #     return super(ContactView, self).form_valid(form)

# class EvalView(View):

#     # template_name = "aptitude/evaluacion_view.html"

#     def get_context_data(self, **kwargs):
#         context = super(EvalView, self).get_context_data(**kwargs)
#         evaluacion = Evaluacion.objects.get(pk=kwargs['eval_id'])
#         if evaluacion is None:
#             messages.error(self.request, 'Evaluation not found')
#             return reverse('evaluacion:list')

#     def get(self, request, *args, **kwargs):
#         form = EvaluacionForm()
#         context = {'form': form}
#         return render(request, 'contact-us.html', context)

#     def post(self, request, *args, **kwargs):
#         form = EvaluacionForm(data=request.POST)
#         if form.is_valid():
#             self.send_mail(form.cleaned_data)
#             form = EvaluacionForm()
#             return render(request, 'aptitude/evaluacion_view.html', {'form': form})
#         return render(request, 'aptitude/evaluacion_view.html', {'form': form})

#     def send_mail(self, valid_data):
#         # Send mail logic
#         print(valid_data)
#         pass



class CuestionarioListView(LoginRequiredMixin, ListView):
    model = Cuestionario

    def get_queryset(self):
        return Cuestionario.objects.filter()


class FormCreate(LoginRequiredMixin, CreateView):
    model = Cuestionario
    template_name = 'aptitude_crd/form_create.html'
    fields = '__all__'

    def post(self, request):
        result = {"result": "", "error_reason": ""}
        unicode_body = request.body.decode('utf-8')
        dict_post_data = json.loads(unicode_body)
        if len(dict_post_data['questions']) > 0:
            cuestionario = Cuestionario(nombre=dict_post_data['evaluacion_nombre'],
                                  description=dict_post_data['evaluacion_description'])
            cuestionario.save()
            result['result'] = 'Form saved successfully'
            for question_item in dict_post_data['componentes']:
                componente = Componente(componente_nombre=question_item['nombre'],
                                      componente_tipo=question_item['tipo'],
                                      cuestionario=cuestionario)
                componente.save()
                if question_item['type'] == 'mcq_one' or question_item['type'] == 'mcq_many':
                    for choice_item in question_item['options']:
                        choice = Opcion(nombre=choice_item,
                                        componente=componente)
                        choice.save()
        else:
            result['result'] = 'Add a question title'
        return HttpResponse(json.dumps(result))

def prepare_workbook(df, transpose=False):

    # Transpose the DataFrame

    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'APTITUDE-PROXY Step 1'

    if transpose:

        # Add variable names as the first column
        for row_idx, column in enumerate(df.columns, start=1):
            sheet.cell(row=row_idx, column=1, value=column)

        # Write the DataFrame to the worksheet column-wise

        for col_idx, row in enumerate(df.itertuples(index=False), start=2):
            for row_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    else:

       # Write the DataFrame to the worksheet
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Add column headers
        for col_idx, column in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=col_idx, value=column)

    return workbook

@login_required
def download_proxy(request):

    fecha = datetime.now().strftime("%Y-%m-%d")
    filename = "database_fragilidad_{}.xlsx".format(fecha)

    # excel_file = IO()
    # xlwriter = pd.ExcelWriter(excel_file, engine='xlsxwriter')

    from django.db import connections

    # url_host = request.META.get("HTTP_HOST")
    db_name = 'default'
    # if url_host.startswith("cat."):
    #     db_name = "cat"

    variable_names = [
        "identification_nb",
        "sex",
        "month_birth",
        "year_birth",
        "zip_code",
        "city",
        "prevention_research",
        "prevention_share",
        "date_step1",
        "registration",
        "memory_problem",
        "today_date",
        "place_question",
        "recall_test",
        "five_chair_complet",
        "weight_loss",
        "appetite_loss",
        "problem_eye",
        "right_ear",
        "left_ear",
        "decrease_hearing",
        "depress",
        "little_interest",
        "cognition_alert",
        "nutrition_alert",
        "vision_alert",
        "hearing_alert",
        "psychologie_alert",
        "locomotion_alert",
        "nb_alerts",
        "tool"
    ]


    db_conn = connections[db_name]
    df = pd.read_sql('''select
                   codigo as identification_nb,
                   sexo as sex,
                   strftime('%m', fecha_nacimiento) AS month_birth,
                   strftime('%Y', fecha_nacimiento) AS year_birth,
                   cp as zip_code, 
                   ciudad as city,
                   TRUE as prevention_research,
                   TRUE as prevention_share,
                   v.fecha as date_step1,
                   'professional' as registration,
                   'NAV Database' as tool,
                   'MD' as right_ear,
                    'MD' as  left_ear,
                    'MD' as today_date,
                    'MD' as place_question,
                   v.id as visita
                   from aptitude_paciente p
				   inner join aptitude_visita v on v.paciente_id = p.id''', db_conn) #.to_excel(xlwriter, 'Datos pacientes')

    evaluaciones = Evaluacion.objects.filter(cuestionario=18).filter(completada=1)

    if evaluaciones:

        datos = []
        for eval in evaluaciones:

            

            e = {'identification_nb': eval.visita.paciente.codigo,
                'visita': eval.visita.id}
            e.update(dict([(p.componente.nombre.strip('_'), p.valor) for p in eval.preguntas.all() if not p.componente.nombre.endswith('_navarra')]))
            

            nb_alerts = eval.preguntas.filter(componente__nombre__endswith='_alert').filter(valor='SI').count() 
            e.update({"nb_alerts": nb_alerts})
            
            datos.append(e)
        #preguntas = pd.read_json(json.dumps(datos))


        # Convert the dictionary to a JSON string
        json_data = json.dumps(datos)

        # Wrap the JSON string in a StringIO object
        json_file = StringIO(json_data)

        # Read the JSON into a pandas DataFrame
        preguntas = pd.read_json(json_file)

        df = df.merge(preguntas, on=['identification_nb','visita'])

        df = df.drop('visita', axis=1)

        df = df[variable_names]
    else:
        df = pd.DataFrame(columns=variable_names)

    # Create a new Excel workbook
    workbook = prepare_workbook(df, transpose=True)

    # Save the workbook to an in-memory buffer
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Create an HTTP response with the Excel file as an attachment
    response = HttpResponse(
        excel_file,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def download_db(request):

    fecha = datetime.now().strftime("%Y-%m-%d")
    filename = "database_fragilidad_{}.xlsx".format(fecha)

    excel_file = IO()
    xlwriter = pd.ExcelWriter(excel_file, engine='xlsxwriter')

    from django.db import connections

    url_host = request.META.get("HTTP_HOST")
    db_name = 'default'
    if url_host.startswith("cat."):
        db_name = "cat"

    db_conn = connections[db_name]
    pd.read_sql('''select
                   codigo,
				   v.fecha as fecha_visita,
                   fecha_nacimiento,
                   sexo,
                   internet,
                   nivel_estudios,
                   captacion,
                   territorio,
                   cp,
                   vivienda,
                   convivientes,
                   salud,
                   problemas_num,
                   problemas
                   from aptitude_paciente p
				   inner join aptitude_visita v on v.paciente_id = p.id''', db_conn).to_excel(xlwriter, 'Datos pacientes')

    cuestionarios = Cuestionario.objects.filter(activo=1)
    for c in cuestionarios:

        datos = []

        evaluaciones = Evaluacion.objects.filter(cuestionario=c)
        for evaluacion in evaluaciones:
            e = {'codigo': evaluacion.visita.paciente.codigo,
                 'fecha': evaluacion.visita.fecha.strftime("%Y-%m-%d") }
            e.update(dict([(p.componente.nombre, p.valor) for p in evaluacion.preguntas.all()]))
            datos.append(e)

        pd.read_json(json.dumps(datos)).to_excel(xlwriter, c.nombre[:30])


    xlwriter.save()
    xlwriter.close()

    excel_file.seek(0)

    response = HttpResponse(excel_file.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    return response

@login_required
def download_registro(request):

    fecha = datetime.now().strftime("%Y-%m-%d")
    filename = "registro_efa_{}.xlsx".format(fecha)

    registros = Registro.objects.all()

    datos = [{'nombre': r.nombre,
              'apellidos': r.apellidos,
              'sexo': r.sexo,
              'edad': r.edad,
              'fecha nacimiento': r.fecha_nacimiento.strftime("%Y-%m-%d"),
              'email': r.email,
              'dni': r.dni,
              'direccion': r.direccion,
              'telefono': r.telefono,
              'ciudad': r.ciudad,
              'pais': r.pais,
              'temas': ','.join([tema.titulo for tema in r.temas.all()])} for r in registros]

    excel_file = IO()
    xlwriter = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    pd.read_json(json.dumps(datos)).to_excel(xlwriter, "Registro")
    xlwriter.save()
    xlwriter.close()

    excel_file.seek(0)

    response = HttpResponse(excel_file.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    return response

@login_required
def view_form(request, eval_id):
    cuestionario = Cuestionario.objects.get(id=eval_id)
    componentes = Componente.objects.filter(cuestionario=cuestionario)
    choices = Opcion.objects.filter(componente__in=componentes)
    context = {
        'form': cuestionario,
        'questions': componentes,
        'choices': choices
    }
    return render(request, 'aptitude_crd/form_view.html', context)


# unused in app

from .serializers import UserSerializer, CuestionarioSerializer, ComponenteSerializer

class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer


class CuestionarioViewSet(viewsets.ModelViewSet):
    queryset = Cuestionario.objects.all()
    serializer_class = CuestionarioSerializer


class ComponenteViewSet(viewsets.ModelViewSet):
    queryset = Componente.objects.all()
    serializer_class = ComponenteSerializer
